import os
import asyncio
import base64
import json
import re
import logging
from pathlib import Path
from datetime import datetime, timedelta
from io import BytesIO

import anthropic
from docx import Document
import gspread
from google.oauth2.service_account import Credentials
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import Application, MessageHandler, CommandHandler, filters, ContextTypes

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

TELEGRAM_TOKEN    = os.environ["TELEGRAM_TOKEN"]
ANTHROPIC_API_KEY = os.environ["ANTHROPIC_API_KEY"]
GOOGLE_SHEET_ID      = os.environ["GOOGLE_SHEET_ID"]
GOOGLE_CREDS_JSON    = os.environ["GOOGLE_CREDS_JSON"]
GOOGLE_DRIVE_FOLDER  = os.environ.get("GOOGLE_DRIVE_FOLDER", "")
TELEGRAM_GROUP_ID    = os.environ.get("TELEGRAM_GROUP_ID", "")

HEADERS       = ["ФИО", "Телефон", "Возраст", "Город", "Должности", "Источник", "Опыт работы", "Ссылка", "Кто добавил", "Дата"]
CRM_HEADERS   = ["ФИО", "НОМЕР", "ГОРОД", "ВОЗРАСТ", "ПОЗИЦИЯ", "ИСТОЧНИК"]
DATE_FMT      = "%d.%m.%Y %H:%M"

client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

pending: dict[int, list] = {}

# ── Google Sheets ──────────────────────────────────────────────────────────────

def get_gc():
    creds_data = json.loads(GOOGLE_CREDS_JSON)
    scopes = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds = Credentials.from_service_account_info(creds_data, scopes=scopes)
    return gspread.authorize(creds)

def upload_to_drive(file_bytes: bytes, filename: str, city: str, mime_type: str):
    """Загружает файл на Google Drive в папку Город/YYYY-MM-DD."""
    if not GOOGLE_DRIVE_FOLDER:
        return
    try:
        from googleapiclient.discovery import build
        from googleapiclient.http import MediaIoBaseUpload
        creds_data = json.loads(GOOGLE_CREDS_JSON)
        scopes = ["https://www.googleapis.com/auth/drive"]
        from google.oauth2.service_account import Credentials
        creds = Credentials.from_service_account_info(creds_data, scopes=scopes)
        service = build("drive", "v3", credentials=creds)

        date_str = datetime.now().strftime("%Y-%m-%d")
        city_name = city if city else "Без города"

        def get_or_create_folder(name, parent_id):
            q = f"name='{name}' and '{parent_id}' in parents and mimeType='application/vnd.google-apps.folder' and trashed=false"
            res = service.files().list(
                q=q, fields="files(id)",
                supportsAllDrives=True, includeItemsFromAllDrives=True
            ).execute()
            if res["files"]:
                return res["files"][0]["id"]
            meta = {"name": name, "mimeType": "application/vnd.google-apps.folder", "parents": [parent_id]}
            f = service.files().create(
                body=meta, fields="id",
                supportsAllDrives=True
            ).execute()
            return f["id"]

        city_folder_id = get_or_create_folder(city_name, GOOGLE_DRIVE_FOLDER)
        date_folder_id = get_or_create_folder(date_str, city_folder_id)

        media = MediaIoBaseUpload(BytesIO(file_bytes), mimetype=mime_type)
        service.files().create(
            body={"name": filename, "parents": [date_folder_id]},
            media_body=media,
            fields="id",
            supportsAllDrives=True
        ).execute()
        log.info(f"Uploaded {filename} to Drive: {city_name}/{date_str}")
    except Exception as e:
        log.warning(f"Drive upload error: {e}")

# Кэш тем: город -> topic_id
_topic_cache: dict[str, int] = {}
_topic_lock = asyncio.Lock()

async def get_or_create_topic(bot, city: str) -> int | None:
    """Возвращает ID темы для города, создаёт если нет. Без дублей."""
    if not TELEGRAM_GROUP_ID:
        return None
    group_id = int(TELEGRAM_GROUP_ID)
    city_name = city if city else "Без города"

    # Лок — чтобы параллельные задачи не создавали тему одновременно
    async with _topic_lock:
        # Проверяем кэш
        if city_name in _topic_cache:
            return _topic_cache[city_name]

        # Ищем существующую тему через getForumTopics
        try:
            topics = await bot.get_forum_topics(chat_id=group_id)
            for t in (topics.topics if topics else []):
                if t.name == city_name:
                    _topic_cache[city_name] = t.message_thread_id
                    log.info(f"Found existing topic '{city_name}': {t.message_thread_id}")
                    return t.message_thread_id
        except Exception as e:
            log.warning(f"get_forum_topics error: {e}")

        # Не нашли — создаём новую
        try:
            topic = await bot.create_forum_topic(
                chat_id=group_id,
                name=city_name,
            )
            _topic_cache[city_name] = topic.message_thread_id
            log.info(f"Created topic '{city_name}': {topic.message_thread_id}")
            return topic.message_thread_id
        except Exception as e:
            log.warning(f"Topic create error for '{city_name}': {e}")
            return None

async def send_file_to_group(bot, file_bytes: bytes, filename: str, city: str,
                              fio: str, phone: str, mime_type: str):
    """Отправляет файл в нужную тему группы."""
    if not TELEGRAM_GROUP_ID:
        return
    try:
        group_id = int(TELEGRAM_GROUP_ID)
        topic_id = await get_or_create_topic(bot, city)
        caption = f"👤 {fio}\n📞 {phone}\n🏙 {city if city else '—'}"
        buf = BytesIO(file_bytes)
        kwargs = dict(
            chat_id=group_id,
            document=buf,
            filename=filename,
            caption=caption,
            read_timeout=60,
            write_timeout=60,
            connect_timeout=30,
        )
        if topic_id:
            kwargs["message_thread_id"] = topic_id
        await bot.send_document(**kwargs)
        log.info(f"Sent {filename} to group topic '{city}'")
    except Exception as e:
        log.warning(f"Group send error: {e}")
        # Повторная попытка через 3 секунды
        try:
            await asyncio.sleep(3)
            buf2 = BytesIO(file_bytes)
            kwargs2 = dict(
                chat_id=group_id,
                document=buf2,
                filename=filename,
                caption=caption,
                read_timeout=60,
                write_timeout=60,
            )
            if topic_id:
                kwargs2["message_thread_id"] = topic_id
            await bot.send_document(**kwargs2)
            log.info(f"Retry OK: {filename}")
        except Exception as e2:
            log.warning(f"Retry failed: {e2}")

def get_sheet(name="Резюме"):
    gc = get_gc()
    sh = gc.open_by_key(GOOGLE_SHEET_ID)
    try:
        ws = sh.worksheet(name)
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(name, rows=5000, cols=20)
        if name == "Резюме":
            ws.append_row(HEADERS)
            ws.format("A1:J1", {"textFormat": {"bold": True}})
        elif name == "Экспорты":
            ws.append_row(["chat_id", "last_export"])
    return ws

CITY_ALIASES = {
    # Днепр
    "дніпро": "Днепр", "дніпропетровськ": "Днепр", "днепропетровск": "Днепр", "днепр": "Днепр",
    # Киев
    "київ": "Киев", "киев": "Киев",
    # Харьков
    "харків": "Харьков", "харьков": "Харьков",
    # Кривой Рог
    "кривий ріг": "Кривой Рог", "кривой рог": "Кривой Рог",
    # Запорожье
    "запоріжжя": "Запорожье", "запорожье": "Запорожье",
    # Львов
    "львів": "Львов", "львов": "Львов",
    # Одесса
    "одеса": "Одесса", "одесса": "Одесса",
    # Николаев
    "миколаїв": "Николаев", "николаев": "Николаев", "миколів": "Николаев",
    # Каменское
    "кам'янське": "Каменское", "камянське": "Каменское", "каменское": "Каменское",
    "кам'янске": "Каменское", "камянске": "Каменское",
    # Ивано-Франковск
    "івано-франківськ": "Ивано-Франковск", "івано франківськ": "Ивано-Франковск",
    "ивано-франковск": "Ивано-Франковск", "ивано франковск": "Ивано-Франковск",
    # Другие
    "полтава": "Полтава",
    "вінниця": "Винница", "винница": "Винница",
    "павлоград": "Павлоград",
    "нікополь": "Никополь", "никополь": "Никополь",
    "черкаси": "Черкассы", "черкассы": "Черкассы",
    "суми": "Сумы", "сумы": "Сумы",
    "житомир": "Житомир",
    "рівне": "Ровно", "ровно": "Ровно",
    "луцьк": "Луцк", "луцк": "Луцк",
    "тернопіль": "Тернополь", "тернополь": "Тернополь",
    "херсон": "Херсон",
    "хмельницький": "Хмельницкий", "хмельницкий": "Хмельницкий",
    "чернівці": "Черновцы", "черновцы": "Черновцы",
    "чернігів": "Чернигов", "чернигов": "Чернигов",
}

def normalize_city(city: str) -> str:
    from difflib import get_close_matches
    if not city:
        return ""
    key = city.strip().lower()
    # 1. Точное совпадение в словаре
    if key in CITY_ALIASES:
        return CITY_ALIASES[key]
    # 2. Нечёткий поиск по ключам словаря (cutoff=0.82 — достаточно строго)
    matches = get_close_matches(key, CITY_ALIASES.keys(), n=1, cutoff=0.82)
    if matches:
        return CITY_ALIASES[matches[0]]
    # 3. Нечёткий поиск по значениям (уже нормализованным названиям)
    normalized_values = list(set(CITY_ALIASES.values()))
    values_lower = {v.lower(): v for v in normalized_values}
    matches2 = get_close_matches(key, values_lower.keys(), n=1, cutoff=0.82)
    if matches2:
        return values_lower[matches2[0]]
    # 4. Не нашли — возвращаем как есть с заглавной буквы
    return city.strip().title()

def get_or_create_city_sheet(sh, city: str):
    """Получает или создаёт лист для города."""
    sheet_name = city if city else "Без города"
    try:
        ws = sh.worksheet(sheet_name)
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(sheet_name, rows=5000, cols=len(HEADERS))
        ws.append_row(HEADERS)
        ws.format(f"A1:J1", {"textFormat": {"bold": True},
                              "backgroundColor": {"red": 0.27, "green": 0.51, "blue": 0.71}})
    return ws

def append_rows_batch(rows: list[list]):
    gc = get_gc()
    sh = gc.open_by_key(GOOGLE_SHEET_ID)

    # 1. Главный лист "Резюме" — все записи
    try:
        ws_main = sh.worksheet("Резюме")
    except gspread.WorksheetNotFound:
        ws_main = sh.add_worksheet("Резюме", rows=5000, cols=len(HEADERS))
        ws_main.append_row(HEADERS)
        ws_main.format("A1:J1", {"textFormat": {"bold": True}})
    ws_main.append_rows(rows, value_input_option="USER_ENTERED")

    # 2. Листы по городам
    city_groups: dict[str, list] = {}
    for row in rows:
        raw_city = row[3] if len(row) > 3 else ""
        city = normalize_city(raw_city)
        row[3] = city  # обновляем нормализованный город
        city_groups.setdefault(city, []).append(row)

    for city, city_rows in city_groups.items():
        ws_city = get_or_create_city_sheet(sh, city)
        ws_city.append_rows(city_rows, value_input_option="USER_ENTERED")

def get_last_export(chat_id: int) -> datetime | None:
    """Возвращает время последнего экспорта для данного chat_id."""
    try:
        ws = get_sheet("Экспорты")
        records = ws.get_all_records()
        for r in records:
            if str(r.get("chat_id")) == str(chat_id):
                raw = r.get("last_export", "")
                if raw:
                    return datetime.strptime(raw, DATE_FMT)
    except Exception as e:
        log.warning(f"get_last_export error: {e}")
    return None

def set_last_export(chat_id: int, dt: datetime):
    """Сохраняет время экспорта."""
    try:
        ws = get_sheet("Экспорты")
        records = ws.get_all_values()
        # Ищем строку с этим chat_id
        for i, row in enumerate(records[1:], start=2):
            if row and str(row[0]) == str(chat_id):
                ws.update(f"B{i}", [[dt.strftime(DATE_FMT)]])
                return
        # Не нашли — добавляем новую строку
        ws.append_row([str(chat_id), dt.strftime(DATE_FMT)])
    except Exception as e:
        log.warning(f"set_last_export error: {e}")

def get_unexported_rows(user_name: str):
    """Возвращает (rows, row_numbers, exp_col) — все НЕэкспортированные строки пользователя."""
    ws = get_sheet("Резюме")
    all_rows = ws.get_all_values()
    if len(all_rows) <= 1:
        return [], [], None

    headers = [h.strip() for h in all_rows[0]]
    try:
        who_idx = headers.index("Кто добавил")
    except ValueError:
        who_idx = 8

    # Колонка "Экспортировано" — создаём если нет
    if "Экспортировано" in headers:
        exp_idx = headers.index("Экспортировано")
    else:
        exp_idx = len(headers)
        ws.update_cell(1, exp_idx + 1, "Экспортировано")

    result_rows = []
    result_nums = []
    for i, row in enumerate(all_rows[1:], start=2):
        if len(row) <= who_idx:
            continue
        who = row[who_idx].strip()
        exported = row[exp_idx].strip() if len(row) > exp_idx else ""
        if who.lower() != user_name.strip().lower():
            continue
        if exported:
            continue
        result_rows.append(row)
        result_nums.append(i)
    return result_rows, result_nums, exp_idx + 1  # 1-based col

def mark_exported(row_numbers: list[int], exp_col: int):
    """Ставит отметку 'да' в колонке Экспортировано для указанных строк."""
    if not row_numbers:
        return
    from gspread.cell import Cell
    ws = get_sheet("Резюме")
    cells = [Cell(row=r, col=exp_col, value="да") for r in row_numbers]
    ws.update_cells(cells)

# ── xlsx генерация ─────────────────────────────────────────────────────────────

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "crm_template.xlsx")

def generate_crm_xlsx(rows: list[list]) -> BytesIO:
    """Генерирует xlsx точной копией шаблона CRM."""
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # Удаляем все строки с данными (оставляем только заголовок)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    # Стили — точно как в шаблоне
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    arial_bold = Font(name="Arial", bold=True)
    jetbrains  = Font(name="JetBrains Mono", bold=False)
    white_fill = PatternFill("solid", fgColor="FFFFFFFF")

    # rows колонки: ФИО(0) Телефон(1) Возраст(2) Город(3) Должности(4) Источник(5) Кто(6) Дата(7)
    for r_i, row in enumerate(rows, 2):
        fio      = row[0] if len(row) > 0 else ""
        phone    = str(row[1]).replace(".0", "") if len(row) > 1 else ""
        city     = row[3] if len(row) > 3 else ""
        age      = str(row[2]).replace(".0", "") if len(row) > 2 else ""
        position = row[4] if len(row) > 4 else ""
        source   = row[5] if len(row) > 5 else ""

        vals = [fio, phone, city, age, position, source]
        ws.row_dimensions[r_i].height = 15.75

        for c_i, val in enumerate(vals, 1):
            cell = ws.cell(row=r_i, column=c_i, value=val)
            cell.border = border
            cell.fill = white_fill
            cell.alignment = Alignment(horizontal="left", vertical="bottom")
            if c_i == 6:  # ИСТОЧНИК — JetBrains Mono
                cell.font = jetbrains
            else:
                cell.font = arial_bold
            if c_i == 2:  # НОМЕР — текст, без .0 и E+11
                cell.number_format = "@"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── Claude extraction ──────────────────────────────────────────────────────────

SYSTEM_PROMPT = """Ты извлекаешь данные из резюме с сайтов поиска работы.

Верни ТОЛЬКО JSON без пояснений, markdown и дополнительного текста:
{
  "fio": "Фамилия Имя Отчество на языке оригинала",
  "phone": "380XXXXXXXXX (только цифры, начиная с 380, без пробелов тире скобок плюсов)",
  "age": "только цифра, например 23",
  "city": "город на русском языке — ТОЛЬКО название города без района и области",
  "positions": "должность1, должность2 (на языке оригинала)",
  "source": "Work.ua или Rabota.ua или OLX — определи по логотипу, цветам, стилю сайта. Если не можешь — unknown",
  "experience": "краткое описание опыта работы если есть, например: 2 года продавцом в АТБ, 1 год курьером. Если нет опыта — пустая строка",
  "resume_url": "прямая ссылка на резюме — ищи в адресной строке браузера или внизу страницы. Форматы: work.ua/resumes/XXXXXXX/ или rabota.ua/candidates/XXXXXXX/ или объявление на olx.ua. Если не видно — пустая строка"
}

Правила для телефона:
- 0XX XXX XX XX → 380XXXXXXXXX
- +38 0XX... → 380XXXXXXXXX
- Уже начинается с 380 — оставь как есть
- Только цифры, никаких символов

Правила для города (очень важно — читай внимательно):
- Ищи везде: адресная строка браузера, блок "Місцезнаходження", "Місто", "Город", под фото профиля, рядом с картой
- На OLX: город написан в блоке "Місцезнаходження" — читай ТОЛЬКО город, без района (например "Дніпро" а не "Дніпровський район")
- На Work.ua и Rabota.ua: город в блоке "Місто проживання" или "Город"
- Нормализуй на русский: Дніпро/Дніпропетровськ/Днепропетровск → Днепр, Харків → Харьков, Київ → Киев, Кривий Ріг → Кривой Рог, Запоріжжя → Запорожье, Львів → Львов, Одеса → Одесса, Миколаїв → Николаев, Кам'янське/Кам'янск → Каменское, Новомосковськ → Новомосковск, Павлоград → Павлоград, Вінниця → Винница, Івано-Франківськ → Ивано-Франковск, Черкаси → Черкассы, Суми → Сумы
- Если видишь район или область без города — пустая строка
- Никогда не пиши область (Дніпропетровська область → просто Днепр)

Если поле не найдено — пустая строка "".
"""

def _parse(raw: str) -> dict:
    raw = re.sub(r"^```json|^```|```$", "", raw, flags=re.MULTILINE).strip()
    return json.loads(raw)

def extract_from_image(image_bytes: bytes, mime: str = "image/jpeg") -> dict:
    b64 = base64.standard_b64encode(image_bytes).decode()
    msg = client.messages.create(
        model="claude-haiku-4-5-20251001", max_tokens=800, system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": [
            {"type": "image", "source": {"type": "base64", "media_type": mime, "data": b64}},
            {"type": "text", "text": "Извлеки данные из этого скриншота резюме."}
        ]}],
    )
    return _parse(msg.content[0].text)

def extract_from_pdf(pdf_bytes: bytes) -> dict:
    b64 = base64.standard_b64encode(pdf_bytes).decode()
    msg = client.messages.create(
        model="claude-haiku-4-5-20251001", max_tokens=800, system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": [
            {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": b64}},
            {"type": "text", "text": "Извлеки данные из этого резюме."}
        ]}],
    )
    return _parse(msg.content[0].text)

def extract_from_text(text: str) -> dict:
    msg = client.messages.create(
        model="claude-haiku-4-5-20251001", max_tokens=800, system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": f"Резюме:\n{text}"}],
    )
    return _parse(msg.content[0].text)

def extract_from_docx(docx_bytes: bytes) -> dict:
    doc = Document(BytesIO(docx_bytes))
    text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    return extract_from_text(text)

# ── Process single item ────────────────────────────────────────────────────────

async def process_item(item: dict, bot) -> dict | None:
    try:
        kind = item["kind"]
        if kind == "photo":
            file = await bot.get_file(item["file_id"])
            raw = bytes(await file.download_as_bytearray())
            data = extract_from_image(raw, "image/jpeg")
        elif kind == "image_doc":
            file = await bot.get_file(item["file_id"])
            raw = bytes(await file.download_as_bytearray())
            mime = "image/png" if item["fname"].endswith(".png") else "image/jpeg"
            data = extract_from_image(raw, mime)
        elif kind == "pdf":
            file = await bot.get_file(item["file_id"])
            raw = bytes(await file.download_as_bytearray())
            data = extract_from_pdf(raw)
            data["_raw"] = raw
            data["_fname"] = item.get("fname", "resume.pdf")
            data["_mime"] = "application/pdf"
        elif kind == "docx":
            file = await bot.get_file(item["file_id"])
            raw = bytes(await file.download_as_bytearray())
            data = extract_from_docx(raw)
            data["_raw"] = raw
            data["_fname"] = item.get("fname", "resume.docx")
            data["_mime"] = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        elif kind == "text":
            data = extract_from_text(item["text"])
        else:
            return None
        data["who"] = item.get("who", "")
        # Отправляем файл в Telegram группу по теме-городу
        if "_raw" in data:
            city  = data.get("city", "")
            fio   = data.get("fio", "")
            phone = data.get("phone", "")
            raw   = data.pop("_raw")
            fname = data.pop("_fname")
            mime  = data.pop("_mime")
            await send_file_to_group(bot, raw, fname, city, fio, phone, mime)
        return data
    except Exception as e:
        log.exception(f"Error processing item: {e}")
        return None

# ── Batch processor ────────────────────────────────────────────────────────────

async def safe_edit(msg, text: str):
    """Редактирует сообщение — если нельзя, отправляет новое."""
    try:
        await msg.edit_text(text)
    except Exception:
        try:
            await msg.reply_text(text)
        except Exception:
            pass

async def run_batch(chat_id: int, status_msg, bot):
    items = pending.pop(chat_id, [])
    if not items:
        return

    total = len(items)
    await safe_edit(status_msg, f"⏳ Обрабатываю {total} резюме параллельно...")

    tasks = [process_item(item, bot) for item in items]
    results = await asyncio.gather(*tasks)

    ok, errors = [], []
    rows = []
    now = datetime.now().strftime(DATE_FMT)

    for data in results:
        if data:
            ok.append(data)
            rows.append([
                data.get("fio", ""),
                data.get("phone", ""),
                data.get("age", ""),
                data.get("city", ""),
                data.get("positions", ""),
                data.get("source", ""),
                data.get("experience", ""),
                data.get("resume_url", ""),
                data.get("who", ""),
                now,
            ])
        else:
            errors.append(1)

    if rows:
        try:
            append_rows_batch(rows)
        except Exception as e:
            log.exception(f"Sheet write error: {e}")
            await safe_edit(status_msg, f"❌ Ошибка записи в таблицу: {e}")
            return

    report = f"✅ Готово! Обработано {len(ok)}/{total}"
    if errors:
        report += f"\n❌ Ошибок: {len(errors)}"
    if ok:
        report += "\n\n📋 Добавлено в таблицу:"
        for d in ok[:5]:
            report += (
                f"\n• {d.get('fio') or '—'} | {d.get('phone') or '—'}"
                f" | {d.get('age') or '—'} | {d.get('city') or '—'}"
                f"\n  💼 {d.get('positions') or '—'} | 🌐 {d.get('source') or '—'}"
            )
        if len(ok) > 5:
            report += f"\n\n  ...и ещё {len(ok)-5}"
    report += "\n\n📤 Когда готов выгрузить в CRM — нажми /export"

    await safe_edit(status_msg, report)

# ── Export handler ─────────────────────────────────────────────────────────────

async def cmd_export(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    msg = await update.message.reply_text("⏳ Формирую файл для CRM...")

    try:
        user_name = get_user_name(update)
        log.info(f"EXPORT: user_name='{user_name}', chat_id={chat_id}")
        rows, row_nums, exp_col = get_unexported_rows(user_name)
        log.info(f"EXPORT: found {len(rows)} unexported rows")

        if not rows:
            await safe_edit(msg, f"📭 Нет новых резюме для экспорта.\n\n🔍 Ищу как: {user_name}")
            return

        buf = generate_crm_xlsx(rows)
        now = datetime.now()
        fname = f"crm_import_{now.strftime('%d%m%Y_%H%M')}.xlsx"
        caption = f"📊 Экспорт для CRM\nКонтактов: {len(rows)}"

        await ctx.bot.send_document(
            chat_id=chat_id,
            document=buf,
            filename=fname,
            caption=caption,
        )
        # Помечаем как экспортированные ТОЛЬКО после успешной отправки
        mark_exported(row_nums, exp_col)
        await msg.delete()

    except Exception as e:
        log.exception(f"Export error: {e}")
        await safe_edit(msg, f"❌ Ошибка при экспорте: {e}")

# ── Telegram handlers ──────────────────────────────────────────────────────────

def get_user_name(update: Update) -> str:
    u = update.effective_user
    return u.full_name or u.username or str(u.id)

PROCESS_BTN = "✅ Обработать резюме"
EXPORT_BTN   = "📤 Экспорт в CRM"

def main_keyboard(count: int = 0) -> ReplyKeyboardMarkup:
    """Клавиатура с кнопками. Показывает счётчик если что-то накоплено."""
    process_label = f"✅ Обработать резюме ({count} шт.)" if count > 0 else "✅ Обработать резюме"
    return ReplyKeyboardMarkup(
        [[KeyboardButton(process_label)],
         [KeyboardButton(EXPORT_BTN)]],
        resize_keyboard=True
    )

async def handle_photo(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    photo = update.message.photo[-1]
    pending.setdefault(chat_id, []).append({
        "kind": "photo", "file_id": photo.file_id, "who": get_user_name(update),
    })
    count = len(pending[chat_id])
    await update.message.reply_text(
        f"📥 Накоплено {count} резюме. Когда загрузишь все — нажми кнопку ✅",
        reply_markup=main_keyboard(count)
    )

async def handle_document(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    doc = update.message.document
    fname = (doc.file_name or "").lower()
    mime = doc.mime_type or ""
    ext = Path(fname).suffix

    if ext in (".jpg", ".jpeg") or "jpeg" in mime:
        kind = "image_doc"
    elif ext == ".png" or "png" in mime:
        kind = "image_doc"
    elif ext == ".pdf" or "pdf" in mime:
        kind = "pdf"
    elif ext in (".docx", ".doc") or "word" in mime or "officedocument" in mime:
        kind = "docx"
    else:
        await update.message.reply_text(f"⚠️ Формат {ext} не поддерживается. Жду: jpg, png, pdf, docx.")
        return

    pending.setdefault(chat_id, []).append({
        "kind": kind, "file_id": doc.file_id, "fname": fname, "who": get_user_name(update),
    })
    count = len(pending[chat_id])
    await update.message.reply_text(
        f"📥 Накоплено {count} резюме. Когда загрузишь все — нажми кнопку ✅",
        reply_markup=main_keyboard(count)
    )

async def handle_process_btn(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Кнопка Обработать — запускает батч вручную."""
    chat_id = update.effective_chat.id
    count = len(pending.get(chat_id, []))
    if count == 0:
        await update.message.reply_text(
            "📭 Нет накопленных резюме. Сначала загрузи файлы или фото.",
            reply_markup=main_keyboard(0)
        )
        return
    status_msg = await update.message.reply_text(
        f"⏳ Запускаю обработку {count} резюме...",
        reply_markup=main_keyboard(0)
    )
    await run_batch(chat_id, status_msg, ctx.bot)

async def handle_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    # Кнопки обрабатываем отдельно
    if EXPORT_BTN in text:
        await cmd_export(update, ctx)
        return
    if len(text) < 30:
        return
    chat_id = update.effective_chat.id
    pending.setdefault(chat_id, []).append({
        "kind": "text", "text": text, "who": get_user_name(update),
    })
    count = len(pending[chat_id])
    await update.message.reply_text(
        f"📥 Накоплено {count} резюме. Когда загрузишь все — нажми кнопку ✅",
        reply_markup=main_keyboard(count)
    )

async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    name = update.effective_user.first_name or ""
    await update.message.reply_text(
        f"👋 Привет, {name}!\n\n"
        "Загружай резюме — фото, jpg, png, pdf, docx.\n"
        "Когда загрузишь все — нажми ✅ Обработать.\n\n"
        "📤 Экспорт в CRM — выгрузить xlsx файл для импорта.",
        reply_markup=main_keyboard(0)
    )

# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    app = Application.builder().token(TELEGRAM_TOKEN).build()
    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("export", cmd_export))
    app.add_handler(MessageHandler(filters.Regex(r"✅ Обработать"), handle_process_btn))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    log.info("Bot started")
    app.run_polling()

if __name__ == "__main__":
    main()
